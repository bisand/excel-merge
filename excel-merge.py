from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font

wbIn = load_workbook(filename='amex.xlsx', read_only=True)
wbOut = Workbook()
wsOut = wbOut.active

rowOut = 1
parseHeader = True
colCount = 0

for sheetnameIn in wbIn.sheetnames:
    print("Parsing sheet: " + sheetnameIn)
    wsIn = wbIn[sheetnameIn]
    rowIdxIn = 0
    # font = Font(color=)
    for rowIn in wsIn.rows:
        rowIdxIn = rowIdxIn + 1
        if(not parseHeader and rowIdxIn == 1):
            continue
        colOut = 1
        if(rowIn[0].value is None or rowIn[0].value.strip() == ""):
            break
        for cellIn in rowIn:
            if(parseHeader and (cellIn.value is None or cellIn.value.strip() == "")):
                break
            if(not parseHeader and colOut > colCount):
                break
            wsOut.cell(column=colOut, row=rowOut, value=cellIn.value)
            colOut = colOut + 1
            if(parseHeader):
                colCount = colCount + 1
        rowOut = rowOut + 1
        if(parseHeader):
            parseHeader = False

wbOut.save("output.xlsx")
